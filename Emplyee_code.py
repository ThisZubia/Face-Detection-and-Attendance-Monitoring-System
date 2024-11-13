import tkinter as tk
from tkinter import ttk, messagebox, Toplevel
from tkcalendar import DateEntry
import openpyxl  # Import openpyxl for Excel handling
import re  # For email validation
import os
import cv2  # OpenCV for face detection
import time

class EmployeeManagementSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("Employee Management System")
        self.root.geometry("600x550")  # Increased width to accommodate CNIC field

        # Initialize required attributes
        self.excel_filename = "employee_data.xlsx"
        self.employees = {}  # Store employee data

        # Configure grid layout for the main window
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=0)  # Left side frame
        self.root.grid_columnconfigure(1, weight=1)  # Right side frame

        # Frame for Employee Information with bold title (Left Side)
        self.frame_info = tk.LabelFrame(self.root, text="Employee Information", font=("Arial", 12, "bold"), padx=10, pady=10)
        self.frame_info.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

        # Labels and Entry widgets for Employee Information
        self.create_widgets()

        # Create employee list (Treeview to display employee data)
        self.create_employee_list()

        # Load Excel file (create a new one if not exists)
        self.create_or_load_excel()

    def create_widgets(self):
        """Creates all the input fields and buttons for employee information."""
        # Labels
        labels = ["Department:", "Job Title:", "Employee ID:", "Name:", "Gender:", "DOB:", "Phone No:", "Email:", "Address:", "CNIC:"]
        for i, label in enumerate(labels):
            tk.Label(self.frame_info, text=label).grid(row=i, column=0, sticky="w", pady=5)

        # Comboboxes and Entry widgets
        self.department_cb = ttk.Combobox(
            self.frame_info,
            values=["HR", "IT", "Sales", "Finance", "Marketing", "Logistics", "Operations", "Support"],
            width=18
        )
        self.job_title_cb = ttk.Combobox(self.frame_info, values=["Manager", "Engineer", "Analyst", "Clerk"], width=18)
        self.employee_id_entry = tk.Entry(self.frame_info, width=20)
        self.employee_id_entry.bind("<FocusOut>", self.validate_employee_id)
        
        # Employee Name Entry with validation to prevent numeric input
        vcmd = (self.root.register(self.validate_name_input), '%S', '%P')  # Register the validation function
        self.employee_name_entry = tk.Entry(self.frame_info, width=20, validate="key", validatecommand=vcmd)
        
        self.gender_cb = ttk.Combobox(self.frame_info, values=["Male", "Female", "Other"], width=18)
        
        # Date Picker for DOB
        self.dob_entry = DateEntry(self.frame_info, width=20, date_pattern="yyyy-mm-dd")
        
        # Restrict phone number entry to numeric input only and 11 digits
        self.phone_entry = tk.Entry(self.frame_info, width=20)
        self.phone_entry.bind("<FocusOut>", self.validate_phone_entry)

        self.email_entry = tk.Entry(self.frame_info, width=20)
        self.email_entry.bind("<FocusOut>", self.validate_email)

        self.address_entry = tk.Entry(self.frame_info, width=20)

        # CNIC entry (must be in proper format)
        self.cnic_entry = tk.Entry(self.frame_info, width=20)
        self.cnic_entry.bind("<FocusOut>", self.validate_cnic_entry)

        # Grid placement of widgets
        widgets = [self.department_cb, self.job_title_cb, self.employee_id_entry, self.employee_name_entry, 
                   self.gender_cb, self.dob_entry, self.phone_entry, self.email_entry, self.address_entry, self.cnic_entry]
        for i, widget in enumerate(widgets):
            widget.grid(row=i, column=1, padx=10, pady=5)

        # Buttons for actions (placed in a single line)
        self.btn_frame = tk.Frame(self.frame_info)
        self.btn_frame.grid(row=10, column=0, columnspan=2, pady=10)

        self.save_button = tk.Button(self.btn_frame, text="Save", command=self.save_info, width=10)
        self.delete_button = tk.Button(self.btn_frame, text="Delete", command=self.delete_info, width=10)
        self.capture_button = tk.Button(self.btn_frame, text="Capture", command=self.capture_images, width=10, state=tk.DISABLED)

        # Grid placement of buttons in one row
        self.save_button.grid(row=0, column=0, padx=5)
        self.delete_button.grid(row=0, column=1, padx=5)
        self.capture_button.grid(row=0, column=2, padx=5)  # Placed in the same row

        # Bind validation functions to input fields to enable the Capture button
        self.employee_name_entry.bind("<KeyRelease>", self.enable_capture_button)
        self.employee_id_entry.bind("<KeyRelease>", self.enable_capture_button)
        self.phone_entry.bind("<KeyRelease>", self.enable_capture_button)
        self.email_entry.bind("<KeyRelease>", self.enable_capture_button)
        self.cnic_entry.bind("<KeyRelease>", self.enable_capture_button)

    def create_employee_list(self):
        """Create the employee list (Treeview widget) to display the saved employee data."""
        self.tree_frame = tk.Frame(self.root)
        self.tree_frame.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")

        # Create treeview widget
        self.tree = ttk.Treeview(self.tree_frame, columns=("Employee ID", "Name", "Department", "Job Title", "CNIC"), show="headings")
        self.tree.pack(fill=tk.BOTH, expand=True)

        # Define columns and headings
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col)

        # Load employee data from the Excel file
        self.load_employee_data()

    def load_employee_data(self):
        """Load employee data from Excel file and populate the Treeview."""
        if os.path.exists(self.excel_filename):
            workbook = openpyxl.load_workbook(self.excel_filename)
            sheet = workbook.active

            # Populate the Treeview with employee data
            for row in sheet.iter_rows(min_row=2):
                self.tree.insert('', 'end', values=(row[0].value, row[1].value, row[2].value, row[3].value, row[7].value))

    def create_or_load_excel(self):
        """Create or load the Excel file."""
        if not os.path.exists(self.excel_filename):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Employee ID", "Name", "Department", "Job Title", "Gender", "DOB", "Phone No", "Email", "Address", "CNIC"])
            workbook.save(self.excel_filename)

    def validate_name_input(self, char, value):
        """Validates that only alphabetic characters and spaces are entered for the employee name."""
        if char.isalpha() or char == " ":
            return True
        else:
            messagebox.showerror("Invalid Input", "Employee name can only contain letters and spaces.")
            self.employee_name_entry.delete(0, tk.END)  # Clear the entry box
            return False

    def validate_phone_entry(self, event):
        """Validates that only numeric input is entered in the phone entry field and enforces 11 digits."""
        current_text = self.phone_entry.get()

        # Only validate if there is text in the phone entry field
        if current_text:
            # Allow only numeric input
            filtered_text = ''.join(filter(str.isdigit, current_text))

            # Restrict the length to exactly 11 digits
            if len(filtered_text) > 11:
                filtered_text = filtered_text[:11]

            self.phone_entry.delete(0, tk.END)
            self.phone_entry.insert(0, filtered_text)

            # Ensure the length is exactly 11 digits
            if len(filtered_text) != 11:
                messagebox.showerror("Invalid Phone Number", "Phone number must be exactly 11 digits.")
                self.phone_entry.delete(0, tk.END)  # Clear the entry box
                return False
        return True

    def validate_email(self, event=None):
        """Validate email format. It should contain '@' and '.com'."""
        email = self.email_entry.get()

        # Only validate if there is text in the email entry field
        if email:
            if "@" not in email or ".com" not in email:
                messagebox.showerror("Invalid Email", "Email format must be valid (e.g., example@domain.com).")
                self.email_entry.delete(0, tk.END)  # Clear the entry box
                return False
        return True

    def validate_cnic_entry(self, event=None):
        """Validate CNIC to ensure it's exactly 13 digits."""
        current_text = self.cnic_entry.get()

        if current_text:
            # Allow only numeric input
            filtered_text = ''.join(filter(str.isdigit, current_text))

            # CNIC must be exactly 13 digits
            if len(filtered_text) != 13:
                messagebox.showerror("Invalid CNIC", "CNIC must be exactly 13 digits.")
                self.cnic_entry.delete(0, tk.END)  # Clear the entry box
                return False

            self.cnic_entry.delete(0, tk.END)
            self.cnic_entry.insert(0, filtered_text)
        return True

    def validate_employee_id(self, event):
        """Validates the Employee ID to ensure it is exactly 5 alphanumeric characters, containing both letters and numbers."""
        employee_id = self.employee_id_entry.get()

        # Check if the employee ID is exactly 5 characters long and contains both letters and numbers
        if len(employee_id) == 5 and re.match(r'^(?=.[a-zA-Z])(?=.\d)[a-zA-Z\d]{5}$', employee_id):
            return True
        else:
            messagebox.showerror("Invalid Employee ID", "Employee ID must be exactly 5 alphanumeric characters containing both letters and numbers.")
            self.employee_id_entry.delete(0, tk.END)  # Clear the entry box
            return False

    def enable_capture_button(self, event=None):
        """Enable the Capture button only when required fields are filled."""
        if self.employee_name_entry.get() and self.employee_id_entry.get() and \
           self.phone_entry.get() and self.email_entry.get() and self.cnic_entry.get():
            self.capture_button.config(state=tk.NORMAL)
        else:
            self.capture_button.config(state=tk.DISABLED)

    def capture_images(self):
        """Capture images using the webcam and save them in a folder named by the employee ID."""
        employee_id = self.employee_id_entry.get()

        # Create a folder named after the employee ID to save images
        save_path = f"employee_data/{employee_id}"
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        # Open the webcam (ID 0 by default)
        cap = cv2.VideoCapture(0)

        # Capture 10 images automatically
        for img_count in range(10):
            ret, frame = cap.read()
            if not ret:
                break

            # Display the captured frame in a window
            cv2.imshow("Capturing Images", frame)

            # Save the captured image
            image_name = f"{save_path}/{employee_id}_{img_count + 1}.jpg"
            cv2.imwrite(image_name, frame)  # Save the image

            # Wait for 1 second before capturing the next image
            cv2.waitKey(1000)

        # Release the webcam and close all OpenCV windows
        cap.release()
        cv2.destroyAllWindows()

        # Inform the user that the capture is complete
        messagebox.showinfo("Capture Complete", "10 images have been captured and saved.")

    def save_info(self):
        """Save the employee information to the Excel file."""
        # Ensure images are saved before saving data
        employee_id = self.employee_id_entry.get()
        image_folder = f"employee_data/{employee_id}"

        if not os.path.exists(image_folder) or len(os.listdir(image_folder)) != 10:
            messagebox.showwarning("No Images Captured", "Please ensure 10 images are captured and saved.")
            return

        # Get employee details from input fields
        name = self.employee_name_entry.get()
        department = self.department_cb.get()
        job_title = self.job_title_cb.get()
        gender = self.gender_cb.get()
        dob = self.dob_entry.get_date() if self.dob_entry.get_date() else ''
        phone = self.phone_entry.get()
        email = self.email_entry.get()
        address = self.address_entry.get()
        cnic = self.cnic_entry.get()

        # Validate required fields
        if not (employee_id and name and department and job_title and phone and email and cnic):
            messagebox.showwarning("Incomplete Data", "Please fill all the required fields.")
            return

        # Check if the Employee ID already exists in the system
        if self.check_existing_data(employee_id):
            messagebox.showwarning("Duplicate Data", "Employee ID already exists.")
            return

        # Create or load the Excel file
        if not os.path.exists(self.excel_filename):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Employee ID", "Name", "Department", "Job Title", "Gender", "DOB", "Phone No", "Email", "Address", "CNIC"])
        else:
            workbook = openpyxl.load_workbook(self.excel_filename)
            sheet = workbook.active

        # Add the employee data to the sheet
        sheet.append([employee_id, name, department, job_title, gender, dob, phone, email, address, cnic])

        # Save the workbook
        workbook.save(self.excel_filename)

        # Update the treeview with the new employee data
        self.tree.insert('', 'end', values=(employee_id, name, department, job_title, cnic))

        # Clear input fields
        self.clear_inputs()

    def check_existing_data(self, employee_id):
        """Check if Employee ID already exists in the Excel file."""
        if os.path.exists(self.excel_filename):
            workbook = openpyxl.load_workbook(self.excel_filename)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if row[0].value == employee_id:
                    return True
        return False

    def delete_info(self):
        """Delete the selected employee from the Excel file and treeview."""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select an employee to delete.")
            return

        # Get employee ID to delete
        employee_id = self.tree.item(selected_item)["values"][0]

        # Load the existing data
        workbook = openpyxl.load_workbook(self.excel_filename)
        sheet = workbook.active

        # Find and delete the row corresponding to the selected employee
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[0].value == employee_id:
                sheet.delete_rows(row[0].row)
                break

        # Save the updated workbook
        workbook.save(self.excel_filename)

        # Delete the row from the treeview
        self.tree.delete(selected_item)

    def clear_inputs(self):
        """Clear input fields."""
        self.employee_id_entry.delete(0, tk.END)
        self.employee_name_entry.delete(0, tk.END)
        self.department_cb.set('')
        self.job_title_cb.set('')
        self.gender_cb.set('')
        self.dob_entry.set_date(None)  # Set date to None instead of empty string
        self.phone_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.address_entry.delete(0, tk.END)
        self.cnic_entry.delete(0, tk.END)


if __name__ == "__main__":
    root = tk.Tk()
    app = EmployeeManagementSystem(root)
    root.mainloop()